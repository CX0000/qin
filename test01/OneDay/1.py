

#class translate():
 #   def __init__(self):
  #      self.conernt={'甘雨':12345678,'钟离':123456789,'凝光':12345678990,'北斗':123456778,'胡桃':13245678,'香菱':13456789}
   # def Precissearch(self):
    #    u=input('请输入你想要搜的内容')
     #   if u in self.conernt:
      #      print(self.conernt.pop(u))
class ExcelOpertions():
    def readExcelData(self,sheetName,x,y):
        import xlrd
        ExcelPath = '../poordata/user.xls'
        OpenFile=xlrd.open_workbook(ExcelPath)
        sheetpage=OpenFile.sheet_by_name(sheetName)
        print(sheetpage.nrows)
        hhhh=[]
        for i in range(1,sheetpage.nrows):
            hhhh.append(sheetpage.cell_value(i,0))
            return hhhh

        #return sheetpage.cell_value(x,y)


print(ExcelOpertions().readExcelData('jjjjj',1,1))

class ExcelOpertions():
    def writeExcelData(self,write_data):
        import openpyxl
        file='../poordata/user.xlsx'
        cx=openpyxl.load_workbook(file)
        sheetPage=cx['jjjjj']
        print(sheetPage.max_row)
        write_row=sheetPage.max_row+1
        sheetPage.cell(write_row,1,write_data)
        cx.save(file)
Excel=ExcelOpertions()
print(Excel.writeExcelData("北斗"))





























































